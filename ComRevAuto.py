import re
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


st.set_page_config(
    page_title="Excel File Comparison Tool",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)


HEADER_KEYWORDS = {
    "",
    "",
    "",
    "",
    "",
    "",
    "",
}

KEY_COLUMN_CANDIDATES = ("PL NUMBER", "Customer No.", "Customer No", "PL")
IGNORE_COMPARE_COLUMNS = {"PL"}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
CHANGE_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
ADDED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
REMOVED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WHITE_BOLD_FONT = Font(color="FFFFFF", bold=True)


def normalize_header(value):
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def dedupe_columns(columns):
    seen = {}
    result = []

    for idx, column in enumerate(columns, start=1):
        name = normalize_header(column) or f"Column {idx}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        result.append(name)

    return result


def header_score(row):
    values = [normalize_header(value).lower() for value in row.dropna()]
    if not values:
        return 0

    score = len(values)
    for value in values:
        if value in HEADER_KEYWORDS or any(keyword in value for keyword in HEADER_KEYWORDS):
            score += 10
    return score


def read_excel_clean(uploaded_file, sheet_name):
    raw_df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None, dtype=object)

    if raw_df.empty:
        return raw_df

    search_rows = min(10, len(raw_df))
    header_row = max(range(search_rows), key=lambda idx: header_score(raw_df.iloc[idx]))

    df = raw_df.iloc[header_row + 1 :].copy()
    df.columns = dedupe_columns(raw_df.iloc[header_row].tolist())
    df = df.dropna(how="all").reset_index(drop=True)

    empty_columns = [column for column in df.columns if df[column].isna().all()]
    return df.drop(columns=empty_columns)


def find_key_column(df):
    normalized_columns = {normalize_header(column).lower(): column for column in df.columns}
    for candidate in KEY_COLUMN_CANDIDATES:
        column = normalized_columns.get(candidate.lower())
        if column is not None:
            return column
    return None


def get_sheet_names(uploaded_file):
    try:
        uploaded_file.seek(0)
        with pd.ExcelFile(uploaded_file) as excel:
            sheet_names = excel.sheet_names
    except Exception:
        sheet_names = []
    finally:
        uploaded_file.seek(0)
    return sheet_names


def normalize_key(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits.lstrip("0") or digits


def comparable_value(value):
    if pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().date().isoformat()
    if hasattr(value, "date") and hasattr(value, "time"):
        return value.date().isoformat()
    if isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
        return round(float(value), 6)
    return re.sub(r"\s+", " ", str(value).strip())


def display_value(value):
    if pd.isna(value):
        return ""
    return value


def build_comparison(original_df, revised_df):
    original_key_col = find_key_column(original_df)
    revised_key_col = find_key_column(revised_df)

    original_work = original_df.copy()
    revised_work = revised_df.copy()

    if original_key_col and revised_key_col:
        original_work["_compare_key"] = original_work[original_key_col].map(normalize_key)
        revised_work["_compare_key"] = revised_work[revised_key_col].map(normalize_key)
    else:
        original_work["_compare_key"] = [str(index + 1) for index in range(len(original_work))]
        revised_work["_compare_key"] = [str(index + 1) for index in range(len(revised_work))]

    original_lookup = {
        row["_compare_key"]: row
        for _, row in original_work.iterrows()
        if row["_compare_key"]
    }
    revised_lookup = {
        row["_compare_key"]: row
        for _, row in revised_work.iterrows()
        if row["_compare_key"]
    }

    original_columns = [column for column in original_df.columns if column not in IGNORE_COMPARE_COLUMNS]
    revised_columns = [column for column in revised_df.columns if column not in IGNORE_COMPARE_COLUMNS]
    common_columns = [column for column in revised_columns if column in original_columns]

    changes = []
    revised_changed_cells = set()
    added_row_numbers = set()
    removed_keys = []

    for revised_index, revised_row in revised_work.iterrows():
        key = revised_row["_compare_key"]
        excel_row = revised_index + 2

        if not key or key not in original_lookup:
            added_row_numbers.add(excel_row)
            changes.append(
                {
                    "Account": display_value(revised_row.get(revised_key_col, key)),
                    "Column": "ROW",
                    "Status": "Added in revised",
                    "Original": "",
                    "Revised": "New row",
                }
            )
            continue

        original_row = original_lookup[key]
        for column in common_columns:
            original_value = original_row.get(column)
            revised_value = revised_row.get(column)

            if comparable_value(original_value) != comparable_value(revised_value):
                changes.append(
                    {
                        "Account": display_value(revised_row.get(revised_key_col, key)),
                        "Column": column,
                        "Status": "Changed",
                        "Original": display_value(original_value),
                        "Revised": display_value(revised_value),
                    }
                )
                revised_changed_cells.add((excel_row, revised_df.columns.get_loc(column) + 1))

    for key, original_row in original_lookup.items():
        if key not in revised_lookup:
            removed_keys.append(key)
            changes.append(
                {
                    "Account": display_value(original_row.get(original_key_col, key)),
                    "Column": "ROW",
                    "Status": "Missing from revised",
                    "Original": "Existing row",
                    "Revised": "",
                }
            )

    added_columns = [column for column in revised_columns if column not in original_columns]
    removed_columns = [column for column in original_columns if column not in revised_columns]

    return {
        "changes_df": pd.DataFrame(changes),
        "changed_cells": revised_changed_cells,
        "added_rows": added_row_numbers,
        "removed_keys": removed_keys,
        "added_columns": added_columns,
        "removed_columns": removed_columns,
        "common_columns": common_columns,
        "original_key_col": original_key_col,
        "revised_key_col": revised_key_col,
    }


def style_worksheet(worksheet, freeze_header=True):
    if freeze_header:
        worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def write_report(original_df, revised_df, comparison):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        revised_df.to_excel(writer, sheet_name="Revised - Highlighted", index=False)
        original_df.to_excel(writer, sheet_name="Original", index=False)

        changes_df = comparison["changes_df"]
        if changes_df.empty:
            changes_df = pd.DataFrame(
                [{"Account": "", "Column": "", "Status": "No changes detected", "Original": "", "Revised": ""}]
            )
        changes_df.to_excel(writer, sheet_name="Changes Summary", index=False)

        for worksheet in writer.book.worksheets:
            style_worksheet(worksheet)

        revised_sheet = writer.sheets["Revised - Highlighted"]

        for row_num, col_num in comparison["changed_cells"]:
            revised_sheet.cell(row=row_num, column=col_num).fill = CHANGE_FILL
            revised_sheet.cell(row=row_num, column=col_num).font = WHITE_BOLD_FONT

        for row_num in comparison["added_rows"]:
            for cell in revised_sheet[row_num]:
                cell.fill = ADDED_FILL

        summary_sheet = writer.sheets["Changes Summary"]
        for row in range(2, summary_sheet.max_row + 1):
            status = summary_sheet.cell(row=row, column=3).value
            fill = None
            if status == "Changed":
                fill = CHANGE_FILL
            elif status == "Added in revised":
                fill = ADDED_FILL
            elif status == "Missing from revised":
                fill = REMOVED_FILL

            if fill:
                for cell in summary_sheet[row]:
                    cell.fill = fill
                    if status == "Changed":
                        cell.font = WHITE_BOLD_FONT

    output.seek(0)
    return output.getvalue()


st.title("📊 Excel File Comparison Tool")
st.markdown("Compare original and revised Excel files to identify changes")

st.sidebar.header("Upload Files")
original_file = st.sidebar.file_uploader(
    "Upload Original Excel File",
    type=["xlsx", "xls"],
    key="original",
)
revised_file = st.sidebar.file_uploader(
    "Upload Revised Excel File",
    type=["xlsx", "xls"],
    key="revised",
)

st.sidebar.subheader("Select sheets to compare")
original_sheet_name = None
revised_sheet_name = None

if original_file is not None:
    original_sheet_names = get_sheet_names(original_file)
    if original_sheet_names:
        default_index = 1 if len(original_sheet_names) > 1 else 0
        original_sheet_name = st.sidebar.selectbox(
            "Original sheet",
            original_sheet_names,
            index=default_index,
            key="original_sheet",
        )

if revised_file is not None:
    revised_sheet_names = get_sheet_names(revised_file)
    if revised_sheet_names:
        revised_sheet_name = st.sidebar.selectbox(
            "Revised sheet",
            revised_sheet_names,
            index=0,
            key="revised_sheet",
        )

if original_file is None or revised_file is None:
    st.info("👈 Please upload both Excel files in the sidebar to start comparing.")
else:
    try:
        with st.spinner("Loading files..."):
            original_df = read_excel_clean(original_file, sheet_name=original_sheet_name or 0)
            revised_df = read_excel_clean(revised_file, sheet_name=revised_sheet_name or 0)
            comparison = build_comparison(original_df, revised_df)

        changes_df = comparison["changes_df"]
        changed_cells_count = len(comparison["changed_cells"])

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Original Rows", len(original_df))
        with col2:
            st.metric("Revised Rows", len(revised_df))
        with col3:
            st.metric("Changed Cells", changed_cells_count)
        with col4:
            st.metric("New Rows", len(comparison["added_rows"]))

        st.caption(
            "Matched rows by account number"
            if comparison["original_key_col"] and comparison["revised_key_col"]
            else "No account-number column found; matched rows by position"
        )

        st.divider()

        original_sheet_label = original_sheet_name if isinstance(original_sheet_name, str) else f"Sheet {original_sheet_name + 1 if isinstance(original_sheet_name, int) else 1}"
        revised_sheet_label = revised_sheet_name if isinstance(revised_sheet_name, str) else f"Sheet {revised_sheet_name + 1 if isinstance(revised_sheet_name, int) else 1}"

        st.subheader(f"📄 Original File ({original_sheet_label})")
        st.dataframe(original_df, use_container_width=True, height=250)

        st.subheader(f"📄 Revised File ({revised_sheet_label})")
        st.dataframe(revised_df, use_container_width=True, height=250)

        st.divider()

        st.subheader("🔍 Comparison Summary")

        col1, col2 = st.columns(2)
        with col1:
            st.write("**Structure Changes:**")
            structure_issues = []

            if comparison["added_columns"]:
                structure_issues.append(f"Columns added: {', '.join(comparison['added_columns'])}")
            if comparison["removed_columns"]:
                structure_issues.append(f"Columns removed: {', '.join(comparison['removed_columns'])}")
            if len(original_df) != len(revised_df):
                structure_issues.append(f"Row count changed: {len(original_df)} -> {len(revised_df)}")

            if structure_issues:
                for issue in structure_issues:
                    st.write(issue)
            else:
                st.write("No structural changes detected")

        with col2:
            st.write("**Data Changes:**")
            if changes_df.empty:
                st.write("No data changes detected")
            else:
                st.write(f"{len(changes_df)} change(s) found")
                st.write(f"{changed_cells_count} revised cell(s) will be highlighted red")

        st.divider()

        st.subheader("📋 Detailed Changes Report")
        if changes_df.empty:
            st.success("✅ No changes detected between the files!")
        else:
            st.dataframe(changes_df, use_container_width=True)

        st.divider()

        st.subheader("💾 Download Comparison Report")
        report_bytes = write_report(original_df, revised_df, comparison)
        st.download_button(
            label="📥 Download Excel Report",
            data=report_bytes,
            file_name="comparison_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"❌ Error processing files: {str(e)}")
        st.info("Please ensure both files are valid Excel files and try again.")
